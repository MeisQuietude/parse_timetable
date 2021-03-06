import datetime
from pathlib import Path

import openpyxl

from config import DATE_FORMAT, DST_DIR, FILENAME_TEMPLATE, GROUP_1, GROUP_2, URI
from utils.tools import fix_xlsx, get_binary, get_next_monday

# TODO: defusedxml


date = get_next_monday()
date_formatted = date.strftime(DATE_FORMAT)  # for fetch file from server by filename

dates = (date + datetime.timedelta(days=i) for i in range(6))
dates_formatted = tuple(
    map(lambda d: d.strftime("%d.%m.%Y"), dates)
)  # for fill date cells in a result file

filename1 = FILENAME_TEMPLATE.replace("GROUP", GROUP_1).replace("DATE", date_formatted)
filename2 = FILENAME_TEMPLATE.replace("GROUP", GROUP_2).replace("DATE", date_formatted)

uri1 = f"{URI}/{filename1}"
uri2 = f"{URI}/{filename2}"

folder = date.strftime("%Y.%m.%d")
path = Path(".", DST_DIR, folder)
path.mkdir(parents=True, exist_ok=True)

path1 = Path(".", DST_DIR, folder, filename1)
path2 = Path(".", DST_DIR, folder, filename2)

try:
    with open(path1, "wb") as f:
        f.write(get_binary(uri1))
    with open(path2, "wb") as f:
        f.write(get_binary(uri2))
except Exception as e:
    input(e)
    exit()

data = []

try:
    wb1 = openpyxl.load_workbook(path1, read_only=True)
except KeyError:
    fix_xlsx(path1)
    wb1 = openpyxl.load_workbook(path1, read_only=True)

ws1 = wb1.active
data.append(
    [
        *[ws1.cell(row, 3).value for row in range(21, 21 + 4)],
        *[ws1.cell(row, 3).value for row in range(28, 28 + 4)],
        *[ws1.cell(row, 3).value for row in range(35, 35 + 4)],
        *[ws1.cell(row, 3).value for row in range(42, 42 + 4)],
        *[ws1.cell(row, 3).value for row in range(49, 49 + 4)],
        *[ws1.cell(row, 3).value for row in range(56, 56 + 4)],
    ]
)
wb1.close()

try:
    wb2 = openpyxl.load_workbook(path2, read_only=True)
except KeyError:
    fix_xlsx(path2)
    wb2 = openpyxl.load_workbook(path2, read_only=True)

ws2 = wb2.active
data.append(
    [
        *[ws2.cell(row, 3).value for row in range(21, 21 + 4)],
        *[ws2.cell(row, 3).value for row in range(28, 28 + 4)],
        *[ws2.cell(row, 3).value for row in range(35, 35 + 4)],
        *[ws2.cell(row, 3).value for row in range(42, 42 + 4)],
        *[ws2.cell(row, 3).value for row in range(49, 49 + 4)],
        *[ws2.cell(row, 3).value for row in range(56, 56 + 4)],
    ]
)
wb2.close()

template = openpyxl.load_workbook(Path("timetable", "Timetable_ISIT.xlsx"))
ws = template.active

# Fill dates
for row, value in zip(range(5, 26, 4), dates_formatted):
    ws.cell(row, 1).value = value

# Fill ISIT-1801
for row, value in zip(range(5, 29), data[0]):
    ws.cell(row, 3).value = value

# Fill ISIT-1802
for row, value in zip(range(5, 29), data[1]):
    ws.cell(row, 4).value = value

template.save(Path(path, "result.xlsx"))
